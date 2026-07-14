#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
harmoniser_depenses.py — moule générique « tableaux larges » (large → long).

Réutilisable par les chantiers financiers 2-5 (dépenses par région / par
activités, SAD, contours financiers). Ce fichier fournit :

  * `depivoter_tableau_large(...)` : dépivote UNE feuille « programme × ligne »
    (colonnes = postes ; lignes = régions/CA/programmes) en table longue, en
    ignorant les colonnes de pourcentage et de total, avec rapport qualité.
  * un driver `main()` pour le **Chantier 3 — Dépenses par région** :
    feuille « Par région » des fichiers annuels de
    `Dépenses par région/dépenses/` (2013-2014 → 2022-2023).

Principe (cf. PLAN §4) :
  1. repérer la vraie ligne d'en-tête (ici : juste au-dessus de « RSS 01 ») ;
  2. dépivoter les colonnes de postes → lignes, une colonne `montant` ;
  3. normaliser les libellés + déduire l'`exercice` du nom de fichier ;
  4. rapport qualité par fichier (postes, régions, total, écart vs GRAND TOTAL).

Zéro chiffre inventé : toute perte de lignes ou tout écart au GRAND TOTAL de la
source est reporté dans `rapport_qualite.csv`.

Sortie Chantier 3 : 20_canonique/depenses_region/depenses_region_long.{parquet,csv}
                    + rapport_qualite.csv
"""
from __future__ import annotations
import re, sys, warnings
from pathlib import Path
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

RACINE = Path(__file__).resolve().parents[1]
SRC_DIR = RACINE / "Dépenses par région" / "dépenses"
OUT = RACINE / "20_canonique" / "depenses_region"
JEU = "depenses_par_region"

# Noms des régions sociosanitaires (RSS) — référence publique MSSS.
RSS_NOMS = {
    1: "Bas-Saint-Laurent", 2: "Saguenay–Lac-Saint-Jean",
    3: "Capitale-Nationale", 4: "Mauricie et Centre-du-Québec",
    5: "Estrie", 6: "Montréal", 7: "Outaouais",
    8: "Abitibi-Témiscamingue", 9: "Côte-Nord", 10: "Nord-du-Québec",
    11: "Gaspésie–Îles-de-la-Madeleine", 12: "Chaudière-Appalaches",
    13: "Laval", 14: "Lanaudière", 15: "Laurentides", 16: "Montérégie",
    17: "Nunavik", 18: "Terres-Cries-de-la-Baie-James",
}

# Colonnes à ne jamais traiter comme un poste de dépense.
COLS_IGNOREES = {"%", "GRAND TOTAL", "TOTAL", ""}

# Table d'alias de programmes (variantes de libellé d'une année à l'autre).
ALIAS_PROGRAMME = {
    "Jeunes en difficultés": "Jeunes en difficulté",
}


def _prog(lib: str) -> str:
    """Canonicalise un libellé de programme via la table d'alias."""
    return ALIAS_PROGRAMME.get(lib, lib)

# Fichiers du Chantier 3 -> exercice (le nom de fichier porte l'année de fin).
FICHIERS_REGION = {
    "Depenses-par-programme-et-par-region-2013-2014.xlsx": "2013-2014",
    "04-Contour-Financier-Depenses-prog-Regions-2014-15.xlsx": "2014-2015",
    "Depenses_par_programme_et_par_regions_1516.xlsx": "2015-2016",
    "depenses-par-programme-et-par-region-1617.xlsx": "2016-2017",
    "depenses-par-programme-et-par-region-1718.xlsx": "2017-2018",
    "depenses-par-programme-et-par-region-1819.xlsx": "2018-2019",
    "depenses-par-programme-et-par-region-1920.xlsx": "2019-2020",
    "depenses-par-programme-et-par-region-2021.xlsx": "2020-2021",
    "depenses-par-programme-et-par-region-2022.xlsx": "2021-2022",
    "depenses-par-programme-et-par-region-2023.xlsx": "2022-2023",
}


# --- Chantier 2 : Dépenses par activités (matrice C/A × programme) ----------
SRC_DIR_ACT = RACINE / "Dépenses par activités"
OUT_ACT = RACINE / "20_canonique" / "depenses_activites"
JEU_ACT = "depenses_par_activites"

# --- Chantier 4 : SAD (soutien à domicile) — par programme × région ---------
SRC_DIR_SAD = RACINE / "SAD"
OUT_SAD = RACINE / "20_canonique" / "depenses_sad"
JEU_SAD = "depenses_sad"


def _norm(s: str) -> str:
    """Normalise un libellé : espaces/retours-ligne compactés."""
    return re.sub(r"\s+", " ", str(s)).strip()


def depivoter_tableau_large(
    path: Path, exercice: str,
    motif_ligne: str = r"^RSS\s*(\d{2})\b",
    sheet: str | None = None,
    jeu: str = JEU,
):
    """Dépivote une feuille « postes en colonnes × entités en lignes ».

    - `motif_ligne` : regex sur la 1re colonne pour reconnaître une ligne de
      données ; le 1er groupe capture le code d'entité (ici le n° de RSS).
    - En-tête = la ligne juste au-dessus de la 1re ligne de données.
    Retourne (lignes: list[dict], rapport: dict).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet is None:
        cand = [s for s in wb.sheetnames if "gion" in s.lower()]
        sheet = cand[0] if cand else wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    rx = re.compile(motif_ligne)

    # 1re ligne de données
    i_data0 = None
    for i, r in enumerate(rows):
        if r and isinstance(r[0], str) and rx.search(r[0].strip()):
            i_data0 = i
            break
    if i_data0 is None:
        return [], {"fichier": path.name, "exercice": exercice, "statut": "AUCUNE LIGNE RSS"}

    header = rows[i_data0 - 1]
    postes = {}  # col_index -> libellé normalisé
    col_grand_total = None
    for j, v in enumerate(header):
        if v in (None, ""):
            continue
        lib = _norm(v)
        if lib.upper() in {"GRAND TOTAL", "TOTAL", "TOTAL QUÉBEC"}:
            col_grand_total = j
            continue
        if lib in COLS_IGNOREES or lib == "%":
            continue
        postes[j] = _prog(lib)

    lignes = []
    n_rss = 0
    somme_montants = 0.0
    ecart_grand_total = 0.0
    non_num = 0
    for r in rows[i_data0:]:
        if not (r and isinstance(r[0], str) and rx.search(r[0].strip())):
            continue
        m = rx.search(r[0].strip())
        rss = int(m.group(1))
        n_rss += 1
        somme_ligne = 0.0
        for j, poste in postes.items():
            val = r[j] if j < len(r) else None
            montant = pd.to_numeric(val, errors="coerce")
            if pd.isna(montant):
                if val not in (None, ""):
                    non_num += 1
                continue
            montant = float(montant)
            somme_ligne += montant
            somme_montants += montant
            lignes.append({
                "jeu": jeu, "exercice": exercice, "rss": rss,
                "region_nom": RSS_NOMS.get(rss, "Inconnu"),
                "programme": poste, "montant": montant,
            })
        # contrôle : somme des postes vs GRAND TOTAL de la source
        if col_grand_total is not None and col_grand_total < len(r):
            gt = pd.to_numeric(r[col_grand_total], errors="coerce")
            if not pd.isna(gt):
                ecart_grand_total += abs(somme_ligne - float(gt))

    rapport = {
        "fichier": path.name, "exercice": exercice, "statut": "OK",
        "n_postes": len(postes), "n_regions": n_rss,
        "n_lignes_long": len(lignes),
        "somme_montants": round(somme_montants, 2),
        "ecart_total_vs_grand_total": round(ecart_grand_total, 2),
        "valeurs_non_numeriques": non_num,
    }
    return lignes, rapport


def depivoter_par_ca(path: Path, exercice: str):
    """Dépivote une matrice « C/A (lignes) × programme (colonnes) ».

    Maquette « Dépenses par activités » : en-tête « C/A | NOM C/A | <postes...>
    | TOTAL ». Le nombre de postes ventilés varie selon l'année (les années
    anciennes ne détaillent que la DI-TSA ; la colonne TOTAL reste le total
    tous programmes du centre d'activités). On dépivote les postes présents ;
    on ignore la colonne TOTAL et la ligne « TOTAL DES PROGRAMMES »."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    ih = None
    for i, r in enumerate(rows):
        if r and isinstance(r[0], str) and r[0].strip() == "C/A":
            ih = i
            break
    if ih is None:
        return [], {"fichier": path.name, "exercice": exercice, "statut": "EN-TÊTE C/A ABSENTE"}

    header = rows[ih]
    postes = {}
    col_total = None
    for j, v in enumerate(header):
        if j < 2 or v in (None, ""):
            continue
        lib = _norm(v)
        if lib.upper().startswith("TOTAL"):
            col_total = j
            continue
        postes[j] = _prog(lib)

    lignes = []
    n_ca = 0
    somme_montants = 0.0
    somme_total_col = 0.0
    non_num = 0
    for r in rows[ih + 1:]:
        nom = r[1] if len(r) > 1 else None
        if nom in (None, ""):
            continue
        if isinstance(nom, str) and nom.strip().upper().startswith("TOTAL"):
            continue  # ligne « TOTAL DES PROGRAMMES »
        n_ca += 1
        ca_code = "" if r[0] in (None, "") else str(r[0]).strip()
        ca_nom = _norm(nom)
        for j, poste in postes.items():
            val = r[j] if j < len(r) else None
            montant = pd.to_numeric(val, errors="coerce")
            if pd.isna(montant):
                if val not in (None, ""):
                    non_num += 1
                continue
            montant = float(montant)
            somme_montants += montant
            lignes.append({
                "jeu": JEU_ACT, "exercice": exercice,
                "centre_activites_code": ca_code, "centre_activites": ca_nom,
                "programme": poste, "montant": montant,
            })
        if col_total is not None and col_total < len(r):
            tv = pd.to_numeric(r[col_total], errors="coerce")
            if not pd.isna(tv):
                somme_total_col += float(tv)

    rapport = {
        "fichier": path.name, "exercice": exercice, "statut": "OK",
        "n_postes": len(postes), "n_centres_activites": n_ca,
        "n_lignes_long": len(lignes),
        "somme_montants_postes": round(somme_montants, 2),
        "somme_colonne_total_source": round(somme_total_col, 2),
        "valeurs_non_numeriques": non_num,
    }
    return lignes, rapport


def run_region():
    OUT.mkdir(parents=True, exist_ok=True)
    toutes = []
    rapports = []
    for fichier, exercice in FICHIERS_REGION.items():
        path = SRC_DIR / fichier
        if not path.exists():
            rapports.append({"fichier": fichier, "exercice": exercice,
                             "statut": "FICHIER ABSENT"})
            continue
        lignes, rap = depivoter_tableau_large(path, exercice)
        toutes.extend(lignes)
        rapports.append(rap)

    long = pd.DataFrame(toutes).sort_values(
        ["exercice", "rss", "programme"]).reset_index(drop=True)
    long["montant"] = long["montant"].round(2)
    rap = pd.DataFrame(rapports)

    long.to_parquet(OUT / "depenses_region_long.parquet", index=False)
    long.to_csv(OUT / "depenses_region_long.csv", index=False, encoding="utf-8")
    rap.to_csv(OUT / "rapport_qualite.csv", index=False, encoding="utf-8")

    print(f"[OK] {len(long):,} lignes -> {OUT/'depenses_region_long.parquet'}")
    print("\n=== Rapport qualité ===")
    print(rap.to_string(index=False))
    ditsa = long[long.programme.str.contains("Déficience intellectuelle")]
    print("\n=== DI-TSA Québec par exercice (contrôle) ===")
    print(ditsa.groupby("exercice")["montant"].sum().apply(lambda v: f"{v:,.0f}").to_string())
    print("\nProgrammes:", sorted(long.programme.unique()))


def run_activites():
    OUT_ACT.mkdir(parents=True, exist_ok=True)
    fichiers = sorted(
        f for f in SRC_DIR_ACT.glob("*.xlsx")
        if re.fullmatch(r"\d{4}-\d{4}", f.stem)
    )
    toutes, rapports = [], []
    for path in fichiers:
        exercice = path.stem
        lignes, rap = depivoter_par_ca(path, exercice)
        toutes.extend(lignes)
        rapports.append(rap)

    long = pd.DataFrame(toutes).sort_values(
        ["exercice", "programme", "centre_activites"]).reset_index(drop=True)
    long["montant"] = long["montant"].round(2)
    rap = pd.DataFrame(rapports)

    long.to_parquet(OUT_ACT / "depenses_activites_long.parquet", index=False)
    long.to_csv(OUT_ACT / "depenses_activites_long.csv", index=False, encoding="utf-8")
    rap.to_csv(OUT_ACT / "rapport_qualite.csv", index=False, encoding="utf-8")

    print(f"[OK] {len(long):,} lignes -> {OUT_ACT/'depenses_activites_long.parquet'}")
    print("\n=== Rapport qualité ===")
    print(rap.to_string(index=False))
    ditsa = long[long.programme.str.contains("Déficience intellectuelle")]
    print("\n=== DI-TSA par C/A — total Québec par exercice (contrôle) ===")
    print(ditsa.groupby("exercice")["montant"].sum().apply(lambda v: f"{v:,.0f}").to_string())


def run_sad():
    """Chantier 4 — SAD par programme × région (maquette récente, dès 2016-2017).

    Réutilise depivoter_tableau_large (colonnes programmes × lignes RSS +
    GRAND TOTAL). Les fichiers antérieurs à 2016-2017 ont une définition
    différente (découpage services/soutien, non ventilé par programme) et sont
    écartés : on ne retient que les feuilles présentant une colonne
    « Déficience intellectuelle et TSA »."""
    OUT_SAD.mkdir(parents=True, exist_ok=True)
    fichiers = sorted(SRC_DIR_SAD.glob("*programme*region*.xlsx")) + \
               sorted(SRC_DIR_SAD.glob("*programme-et-par-region*.xlsx"))
    fichiers = sorted(set(fichiers))
    toutes, rapports = [], []
    for path in fichiers:
        wb = openpyxl.load_workbook(path, data_only=True)
        sh = wb.sheetnames[0]
        # exercice depuis le nom de feuille (suffixe AABB -> 20AA-20BB)
        m = re.search(r"(\d{2})(\d{2})\b", sh)
        if not m:
            rapports.append({"fichier": path.name, "statut": "EXERCICE ILLISIBLE (feuille)"})
            continue
        exercice = f"20{m.group(1)}-20{m.group(2)}"
        # ne garder que la maquette « par programme » (colonne DI-TSA présente)
        rows = list(wb[sh].iter_rows(values_only=True))
        ir = next((i for i, r in enumerate(rows)
                   if r and isinstance(r[0], str) and r[0].strip().startswith("RSS 01")), None)
        header = [(_norm(v) if v not in (None, "") else "") for v in rows[ir - 1]] if ir else []
        if not any("Déficience intellectuelle" in h for h in header):
            rapports.append({"fichier": path.name, "exercice": exercice,
                             "statut": "MAQUETTE ANCIENNE (écartée)"})
            continue
        lignes, rap = depivoter_tableau_large(path, exercice, sheet=sh, jeu=JEU_SAD)
        toutes.extend(lignes)
        rapports.append(rap)

    long = pd.DataFrame(toutes).sort_values(
        ["exercice", "rss", "programme"]).reset_index(drop=True)
    long["montant"] = long["montant"].round(2)
    rap = pd.DataFrame(rapports)

    long.to_parquet(OUT_SAD / "depenses_sad_long.parquet", index=False)
    long.to_csv(OUT_SAD / "depenses_sad_long.csv", index=False, encoding="utf-8")
    rap.to_csv(OUT_SAD / "rapport_qualite.csv", index=False, encoding="utf-8")

    print(f"[OK] {len(long):,} lignes -> {OUT_SAD/'depenses_sad_long.parquet'}")
    print("\n=== Rapport qualité ===")
    print(rap.to_string(index=False))
    ditsa = long[long.programme.str.contains("Déficience intellectuelle")]
    print("\n=== SAD DI-TSA Québec par exercice ===")
    print(ditsa.groupby("exercice")["montant"].sum().apply(lambda v: f"{v:,.0f}").to_string())
    print("\n=== SAD GRAND TOTAL (tous programmes) Québec par exercice ===")
    print(long.groupby("exercice")["montant"].sum().apply(lambda v: f"{v:,.0f}").to_string())


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--jeu", choices=["region", "activites", "sad"], default="region")
    args = ap.parse_args()
    if args.jeu == "region":
        run_region()
    elif args.jeu == "activites":
        run_activites()
    else:
        run_sad()


if __name__ == "__main__":
    sys.exit(main())
