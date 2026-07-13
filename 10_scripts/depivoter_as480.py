#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
depivoter_as480.py
==================
Dépivote les classeurs Excel « pivotés » de l'AS480 pour les millésimes
2010-11 à 2013-14 — les seules années non couvertes par harmoniser_as480.py,
qui les signale à juste titre comme hors périmètre (format structurellement
différent). Ce script comble cette lacune.

Format source (pivoté)
-----------------------
Un classeur = une année. ~115-135 onglets, un onglet par « tranche » de page du
formulaire, les établissements (16 centres jeunesse régionaux) en LIGNES et les
postes de données en COLONNES. Trois familles d'onglets :

  * PxxLyy  (ex. P04L01) : une LIGNE fixe de la page, les colonnes de données
    sont les COLONNES du formulaire  -> colonne = position, ligne = yy.
  * PxxCyy  (ex. P01C01) : une COLONNE fixe de la page, les colonnes de données
    sont les LIGNES du formulaire      -> ligne = position, colonne = yy.
  * Onglets de PLAGE (ex. P03L01À09, P08L03À16, P14L01ÀL05, P17L14&15) : la page
    tient sur une seule colonne, plusieurs lignes empaquetées. Quand le nombre
    de lignes de la plage == le nombre de colonnes de données, on peut mapper
    chaque colonne à sa ligne (colonne = 1). Sinon (pages multi-colonnes
    empaquetées, en-têtes groupés), le mapping (ligne, colonne) numérique n'est
    PAS récupérable de façon fiable -> on laisse ligne/colonne vides mais on
    conserve l'EN-TÊTE textuel, qui reste la clé sémantique sûre.
  * Onglets « pleine page » (Pxx seul, ou Page-xx dans les vieux classeurs) :
    idem, en-tête conservé, ligne/colonne laissés vides.

Choix de conception (honnêteté > fausse précision)
--------------------------------------------------
- Le formulaire AS480 a été RESTRUCTURÉ entre 2013-14 et 2014-15 (pages P06/P07/
  P13 présentes ici, absentes ensuite). Les codes P/L/C de ces années NE peuvent
  donc PAS coïncider avec ceux du pipeline 2014-15+. On ne fabrique aucune
  compatibilité : la clé sémantique fiable est page + feuille + en-tête.
- L'`etablissement_id` (8 chiffres) n'existe pas dans ces sources : seuls le
  numéro de région (rss) et le nom d'établissement sont disponibles.
- Les deux fichiers livrés certaines années sont des DOUBLONS octet pour octet
  (même onglets, mêmes données) : un seul est retenu, l'autre est journalisé.
- Le sous-formulaire (AS480G / AS480A) est lu par onglet dans la cellule RAPPORT
  (« Général » / « Autochtones »).

Sorties (20_canonique/AS480/, SANS écraser l'harmonisé 2014-15+) :
  as480_depivote_2010_2014.csv  (UTF-8-BOM)
  as480_depivote_2010_2014.parquet
  rapport_depivotage.csv

Dépendances : pandas, openpyxl, xlrd (.xls 2010-11), pyarrow (optionnel)
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import pandas as pd

try:
    import openpyxl
except ImportError:  # pragma: no cover
    print("openpyxl requis : pip install openpyxl", file=sys.stderr)
    raise

RACINE = Path(__file__).resolve().parent.parent

# Millésime -> dossier Archives (les 4 années pivotées).
ANNEES = {
    "2010-2011": "Archives/AS480/2010-2111",
    "2011-2012": "Archives/AS480/2011-2012",
    "2012-2013": "Archives/AS480/2012-2013",
    "2013-2014": "Archives/AS480/2013-2014",
}

RAPPORT_VERS_SOUSFORM = {"général": "AS480G", "generale": "AS480G",
                         "autochtones": "AS480A", "autochtone": "AS480A"}


# --------------------------------------------------------------------------- #
# Lecture unifiée xls / xlsx : renvoie (nom_onglet -> liste de lignes)
# --------------------------------------------------------------------------- #
def lire_classeur(chemin: Path) -> dict[str, list[list]]:
    onglets: dict[str, list[list]] = {}
    if chemin.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(chemin))
        for sh in wb.sheets():
            onglets[sh.name] = [
                [sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)
            ]
    else:
        wb = openpyxl.load_workbook(str(chemin), read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            onglets[name] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    return onglets


# --------------------------------------------------------------------------- #
# Classification d'un onglet d'après son nom
# --------------------------------------------------------------------------- #
def classer_onglet(nom: str):
    """Renvoie (page:int, type, meta) ou None si l'onglet n'est pas une page.
    type ∈ {'ligne_fixe','colonne_fixe','plage','pleine'}"""
    n = nom.strip()
    if n.lower() in ("instructions", "instruction"):
        return None
    m = re.match(r"^Page-0*(\d+)$", n, re.I)
    if m:
        return int(m.group(1)), "pleine", None
    m = re.match(r"^P0*(\d+)(.*)$", n, re.I)
    if not m:
        return None
    page = int(m.group(1))
    reste = m.group(2).strip()
    if reste == "":
        return page, "pleine", None
    if re.fullmatch(r"C0*(\d+)", reste, re.I):
        return page, "colonne_fixe", int(re.findall(r"\d+", reste)[0])
    if re.fullmatch(r"L0*(\d+)", reste, re.I):
        return page, "ligne_fixe", int(re.findall(r"\d+", reste)[0])
    # onglet de plage : on tente d'énumérer les lignes citées
    lignes = enumerer_lignes(reste)
    return page, "plage", lignes


def enumerer_lignes(reste: str) -> list[int] | None:
    """Parse 'L01À09', 'L10À14', 'L01ET02', 'L14&15', 'L01ÀL05', 'L02à06',
    'L22À27' -> liste de numéros de ligne. Renvoie None si non interprétable."""
    s = reste.upper().replace(" ", "")
    # plages A..B :  ÀL? ou À ou A (accent variable)
    m = re.match(r"^L0*(\d+)(?:À|A)L?0*(\d+)$", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if a <= b:
            return list(range(a, b + 1))
    # couples 'ET' / '&'
    m = re.match(r"^L0*(\d+)(?:ET|&)L?0*(\d+)$", s)
    if m:
        return [int(m.group(1)), int(m.group(2))]
    # ligne unique résiduelle
    m = re.match(r"^L0*(\d+)$", s)
    if m:
        return [int(m.group(1))]
    return None


def nettoyer_entete(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def to_chiffre(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# --------------------------------------------------------------------------- #
# Extraction d'un onglet -> lignes longues
# --------------------------------------------------------------------------- #
def extraire_onglet(rows, page, typ, meta, exercice, feuille, journal_cells):
    """Renvoie (liste d'enregistrements dict, nb_cellules_source_non_vides)."""
    # 1) ligne d'en-tête = première ligne contenant 'RÉGION'
    h = None
    for i, r in enumerate(rows):
        for c in r:
            if isinstance(c, str) and c.strip().upper().startswith("RÉGION"):
                h = i
                break
        if h is not None:
            break
    if h is None:
        return [], 0

    entete = rows[h]
    # 2) RAPPORT -> sous-formulaire (cherché dans les lignes avant l'en-tête)
    sousform = "AS480G"
    rapport_lbl = "Général"
    for r in rows[:h + 1]:
        for j, c in enumerate(r):
            if isinstance(c, str) and c.strip().upper() == "RAPPORT" and j + 1 < len(r):
                val = str(r[j + 1]).strip()
                rapport_lbl = val
                sousform = RAPPORT_VERS_SOUSFORM.get(val.lower(), "AS480G")
    # 3) colonnes de données = colonnes >=2 avec en-tête non vide
    cols = []  # (idx_dans_ligne, label)
    for j in range(2, len(entete)):
        lbl = nettoyer_entete(entete[j])
        if lbl:
            cols.append((j, lbl))
    if not cols:
        return [], 0

    # 4) mapping (ligne, colonne) selon le type d'onglet
    #    -> renvoie pour chaque position k (0-based) : (ligne|None, colonne|None, mapping)
    n = len(cols)
    lc = []
    if typ == "ligne_fixe":
        for k in range(n):
            lc.append((meta, k + 1, "ligne_fixe"))
    elif typ == "colonne_fixe":
        for k in range(n):
            lc.append((k + 1, meta, "colonne_fixe"))
    elif typ == "plage" and meta is not None and len(meta) == n:
        for k in range(n):
            lc.append((meta[k], 1, "plage_lignes"))
    else:  # plage ambiguë ou pleine page -> en-tête seul
        for k in range(n):
            lc.append((None, None, "entete_seule"))

    # 5) lignes de données
    recs = []
    nb_src = 0
    for r in rows[h + 1:]:
        if not r:
            continue
        reg = r[0]
        nom = r[1] if len(r) > 1 else None
        # arrêt sur 'Total général' / lignes de synthèse / vides
        if reg is None or (isinstance(reg, str) and not reg.strip().isdigit()):
            continue
        try:
            region = int(float(reg))
        except (ValueError, TypeError):
            continue
        nom = nettoyer_entete(nom)
        for k, (j, lbl) in enumerate(cols):
            val = r[j] if j < len(r) else None
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            nb_src += 1
            ligne, colonne, mapping = lc[k]
            plc = None
            if ligne is not None and colonne is not None:
                plc = f"P{page:02d}L{ligne:02d}C{colonne:02d}"
            recs.append({
                "exercice": exercice,
                "exercice_debut": int(exercice[:4]),
                "rapport": rapport_lbl,
                "sous_formulaire": sousform,
                "rss": f"{region:02d}",
                "etablissement_id": "",
                "etablissement_nom": nom,
                "page": f"{page:02d}",
                "souspage": "",
                "ligne": "" if ligne is None else f"{ligne:02d}",
                "colonne": "" if colonne is None else f"{colonne:02d}",
                "plc": plc or "",
                "valeur_saisie": val,
                "chiffre": to_chiffre(val),
                "code_cellule": plc or "",
                "source_feuille": feuille,
                "feuille": feuille,
                "col_index": k + 1,
                "entete": lbl,
                "mapping": mapping,
            })
    return recs, nb_src


# --------------------------------------------------------------------------- #
# Dédup : un fichier par année (les 2 fichiers sont des doublons)
# --------------------------------------------------------------------------- #
def choisir_fichiers(dossier: Path):
    fics = [p for p in sorted(dossier.iterdir())
            if p.suffix.lower() in (".xls", ".xlsx")
            and not p.name.lower().startswith("~$")]
    retenus = sorted(fics)  # ordre déterministe
    if not retenus:
        return None, []
    return retenus[0], retenus[1:]


# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description="Dépivote l'AS480 2010-11 -> 2013-14.")
    ap.add_argument("--root", default=str(RACINE))
    args = ap.parse_args()
    racine = Path(args.root)

    tous = []
    rapport = []
    print("=== AS480 dépivotage (millésimes pivotés 2010-11 -> 2013-14) ===")

    for exercice, sous in ANNEES.items():
        dossier = racine / sous
        if not dossier.is_dir():
            print(f"  [SKIP] {exercice}: dossier absent {dossier}", file=sys.stderr)
            continue
        fic, doublons = choisir_fichiers(dossier)
        if fic is None:
            print(f"  [SKIP] {exercice}: aucun classeur .xls/.xlsx", file=sys.stderr)
            continue
        for d in doublons:
            print(f"  [DOUBLON IGNORÉ] {exercice}: {d.name}")

        onglets = lire_classeur(fic)
        # regrouper par page pour éliminer les onglets 'pleine page' redondants
        par_page = {}
        for nom in onglets:
            cl = classer_onglet(nom)
            if cl is None:
                continue
            page, typ, meta = cl
            par_page.setdefault(page, []).append((nom, typ, meta))

        nb_recs_annee = 0
        nb_src_annee = 0
        pages_vues = set()
        for page, lst in sorted(par_page.items()):
            a_des_tranches = any(t != "pleine" for _, t, _ in lst)
            for nom, typ, meta in lst:
                if a_des_tranches and typ == "pleine":
                    continue  # redondant avec les tranches
                recs, nb_src = extraire_onglet(
                    onglets[nom], page, typ, meta, exercice, nom, None)
                tous.extend(recs)
                nb_recs_annee += len(recs)
                nb_src_annee += nb_src
                pages_vues.add(page)
                rapport.append({
                    "exercice": exercice,
                    "fichier": fic.name,
                    "feuille": nom,
                    "page": f"{page:02d}",
                    "type_onglet": typ,
                    "cellules_source": nb_src,
                    "lignes_ecrites": len(recs),
                    "statut": "ok" if nb_src == len(recs) else "ECART",
                })
        print(f"  {exercice}: {fic.name} -> {nb_recs_annee} cellules, "
              f"{len(pages_vues)} pages, perte={nb_src_annee - nb_recs_annee}")

    if not tous:
        print("Aucune donnée extraite.", file=sys.stderr)
        return 1

    df = pd.DataFrame(tous)
    rap = pd.DataFrame(rapport)

    outdir = racine / "20_canonique" / "AS480"
    outdir.mkdir(parents=True, exist_ok=True)
    csv_path = outdir / "as480_depivote_2010_2014.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    try:
        df.to_parquet(outdir / "as480_depivote_2010_2014.parquet", index=False)
    except Exception as e:  # pyarrow absent
        print(f"  [info] parquet non écrit ({e})")
    rap.to_csv(outdir / "rapport_depivotage.csv", index=False, encoding="utf-8-sig")

    print(f"\nÉcrit : {csv_path}")
    print(f"Total cellules : {len(df)}")
    print("Par sous-formulaire :", df.sous_formulaire.value_counts().to_dict())
    print("Par mapping :", df.mapping.value_counts().to_dict())
    ecarts = rap[rap.statut != "ok"]
    print(f"Onglets en écart (cellule source != ligne écrite) : {len(ecarts)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
